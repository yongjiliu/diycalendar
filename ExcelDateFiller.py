#!/usr/bin/python3
# -*- coding: UTF-8 -*-
import logging
import os
import sys
from copy import copy

import openpyxl
import pandas as pandas
import xlrd
import xlutils
import yaml
from pandas._libs.tslibs.timestamps import Timestamp

from BdDataFetcher import BdDataFetcher


class Config(object):
    def __init__(self, config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as yaml_file:
                data = yaml.load(yaml_file)
                self.excel_path = data['excel_path']
                self.sheet_special = data['sheet_special']
                self.skip_row = data['date_skip_row']
                self.skip_col = data['date_skip_col']
                self.max_length = data['max_length']
                self.holiday_color = data['holiday_color']
                self.workday_color = data['workday_color']
                logging.basicConfig(level=logging.DEBUG,
                            format='%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                            datefmt='%a, %d %b %Y %H:%M:%S')

        except Exception as e:
            logging.error(repr(e))
            sys.exit()

class SpecialDay(object):
    def __init__(self):
        self.is_lunar = False
        self.desc = ''

class ExcelDateFiller(object):
    def __init__(self):
        self.data_fetcher = BdDataFetcher()
        self.target = os.path.splitext(config.excel_path)[0] + '_out' + os.path.splitext(config.excel_path)[-1]
        # try:
        #     shutil.copy(config.excel_path, self.target)
        # except IOError as e:
        #     print("Unable to copy file. %s" % e)
        # except:
        #     print("Unexpected error:", sys.exc_info())
        # self.target_workbook = openpyxl.load_workbook(self.target, data_only=True)

    def fill_date_with_openpyxl(self):
        for sheet in self.target_workbook.worksheets:
            for column_index in range(1, sheet.max_column):
                for row_index in range(1, sheet.max_row):
                    data = sheet.cell(column=column_index, row=row_index)
                    print(data.value)

    def read_with_xlrd(self):
        workbook = xlrd.open_workbook(self.target)
        for sheet in workbook.sheets():
            for column_index in range(0, sheet.ncols):
                for row_index in range(0, sheet.nrows):
                    data = sheet.cell(rowx=row_index, colx=column_index)
                    logging.debug('ctype = {}, value = {}, xf_index = {}'.format(data.ctype, data.value, data.xf_index))

    def write_with_openpyxl(self):
        target_workbook = openpyxl.load_workbook(self.target)
        sheet = target_workbook.get_sheet_by_name('sheet_name')
        sheet.cell(0, 0).value = 'value'
        target_workbook.save()

    def write_with_xlwt(self):
        workbook = xlrd.open_workbook(self.target)
        workbook = xlutils.copy(workbook)
        sheet = workbook.get_sheet(0)
        sheet.write(0, 0, 'value')
        workbook.save()

    def load_special_sheet(self):
        data = {}
        special_sheet = pandas.read_excel(config.excel_path, sheet_name=config.sheet_special, header=0)
        for row_index in range(special_sheet.shape[0]):
            key = special_sheet.iloc[row_index, 0]
            struct_time = pandas.to_datetime(key.timestamp(), unit='s').timetuple()
            key = '{}-{}'.format(struct_time.tm_mon, struct_time.tm_mday)
            value = SpecialDay()
            value.desc = special_sheet.iloc[row_index, 1]
            value.is_lunar = special_sheet.iloc[row_index, 2] == '是'
            data[key] = value
        return data


    def fill_date(self):
        pandas_workbook = pandas.read_excel(config.excel_path, sheet_name=None, skiprows= config.skip_row, keep_default_na=False)
        out_workbook = openpyxl.load_workbook(config.excel_path)

        special_day = self.load_special_sheet()

        day_data = {}
        for sheet_name in pandas_workbook.keys():
            if not sheet_name.endswith('月'):
                continue
            sheet = pandas_workbook.get(sheet_name)
            out_sheet = out_workbook.get_sheet_by_name(sheet_name)

            nrows = sheet.shape[0]
            ncols = sheet.shape[1]
            for row_index in range(nrows):
                for col_index in range(ncols):
                    data = sheet.iloc[row_index, col_index]
                    logging.debug('origin row = {}, col = {}, data = {}'.format(row_index, col_index, data))
                    if type(data) == Timestamp:
                        struct_time = pandas.to_datetime(data.timestamp(), unit='s').timetuple()
                        date = '{}-{}-{}'.format(struct_time.tm_year, struct_time.tm_mon, struct_time.tm_mday)
                        if not day_data.__contains__(date):
                            request_data = self.data_fetcher.request(year_month='{}年{}月'.format(struct_time.tm_year, struct_time.tm_mon))
                            day_data.update(request_data)

                        temp_row = row_index + 2 + config.skip_row
                        temp_col = col_index + 1
                         # weekend color
                        if day_data[date]['cnDay'] == '六' or day_data[date]['cnDay'] == '日':
                            holiday_font = copy(out_sheet.cell(temp_row, temp_col).font)
                            holiday_font.color = config.holiday_color
                            out_sheet.cell(temp_row, temp_col).font = holiday_font
                        # holiday color
                        if day_data[date].__contains__('status'):
                            if day_data[date]['status'] == '1': # 休假
                                holiday_font = copy(out_sheet.cell(temp_row, temp_col).font)
                                holiday_font.color = config.holiday_color
                                out_sheet.cell(temp_row, temp_col).font = holiday_font
                            if day_data[date]['status'] == '2': #班
                                workday_font = copy(out_sheet.cell(temp_row, temp_col).font)
                                workday_font.color = config.workday_color
                                out_sheet.cell(temp_row, temp_col).font = workday_font
                        lunar_date = day_data[date]['lDate']
                        if lunar_date == '初一':
                            lunar_date = '{}月'.format(day_data[date]['lMonth'])
                        # logging.debug('date = {}, value = {}'.format(str(date), lunar_date))
                        temp_content = ''
                        if day_data[date].__contains__('value'):
                            temp_content += day_data[date]['value']
                            if len(temp_content) > config.max_length:
                                temp_content = temp_content[:config.max_length]
                        temp_content += '\n'
                        temp_content += lunar_date
                        # spacial day
                        month_day = day_data[date]['month'] + '-' + day_data[date]['day']
                        if special_day.__contains__(month_day):
                            temp_special_day = special_day.get(month_day)
                            if not temp_special_day.is_lunar:
                                temp_content += '\n'
                                temp_content += temp_special_day.desc

                        lunar_month_day = day_data[date]['lunarMonth'] + '-' + day_data[date]['lunarDate']
                        if special_day.__contains__(lunar_month_day):
                            temp_special_day = special_day.get(lunar_month_day)
                            if temp_special_day.is_lunar:
                                temp_content += '\n'
                                temp_content += temp_special_day.desc

                        temp_row = row_index + 3 + config.skip_row
                        temp_col = col_index + 1
                        out_sheet.cell(temp_row, temp_col).value = temp_content


        out_workbook.save(filename=self.target)
        out_workbook.close()

if __name__ == '__main__':
    config = Config(config_path='config.yaml')
    date_filler = ExcelDateFiller()
    date_filler.fill_date()
